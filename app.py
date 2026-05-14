import io
import base64
import re
import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PILImage

st.set_page_config(
    page_title="Order Sheet Generator · Chhajed Garden",
    page_icon="🌿",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.block-container { padding-top: 2rem; max-width: 1100px; }
h1, h2, h3 { font-family: 'DM Serif Display', serif; }
[data-testid="stFileUploader"] {
    border: 2px dashed #1f5c2e44; border-radius: 12px;
    padding: 1rem; background: #f0f7f1;
}
.app-header {
    background: linear-gradient(135deg, #1f5c2e 0%, #2d7a40 100%);
    color: white; padding: 2rem 2.5rem; border-radius: 16px;
    margin-bottom: 2rem; display: flex; align-items: center; gap: 1.5rem;
}
.app-header h1 { color: white; margin: 0; font-size: 2rem; }
.app-header p  { color: #c8e6c9; margin: 0.3rem 0 0; font-size: 1rem; }
.order-card {
    background: white; border: 1px solid #e0e0e0; border-radius: 12px;
    padding: 1.5rem; margin-bottom: 1.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.order-card h3 { margin-top: 0; color: #1f5c2e; font-size: 1.1rem; }
.product-table { width: 100%; border-collapse: collapse; font-size: 0.88rem; }
.product-table thead tr { background: #1f5c2e; color: white; }
.product-table thead th {
    padding: 10px 12px; text-align: left; font-weight: 600;
    font-size: 0.82rem; letter-spacing: 0.04em; text-transform: uppercase;
}
.product-table tbody tr:nth-child(even) { background: #f8faf8; }
.product-table tbody tr:hover           { background: #e8f5e9; }
.product-table tbody td {
    padding: 9px 12px; border-bottom: 1px solid #eeeeee; vertical-align: middle;
}
.product-table tbody td:first-child { color: #888; font-size: 0.8rem; }
.product-table .img-cell img {
    width: 64px; height: 64px; object-fit: cover;
    border-radius: 6px; border: 1px solid #e0e0e0;
}
.product-table .sku  { color: #666; font-family: monospace; font-size: 0.8rem; }
.product-table .amt  { font-weight: 600; color: #1f5c2e; }
.totals-block {
    background: white; border: 1px solid #e0e0e0;
    border-radius: 12px; overflow: hidden; margin-top: 0;
}
.totals-row {
    display: flex; justify-content: space-between;
    padding: 10px 20px; border-bottom: 1px solid #f0f0f0; font-size: 0.92rem;
}
.totals-row:last-child { border-bottom: none; }
.totals-row.final { background: #1f5c2e; color: white; font-weight: 700; font-size: 1rem; }
.totals-label { font-weight: 500; }
.totals-value { font-weight: 600; }
.info-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0.4rem 2rem; font-size: 0.88rem; }
.info-row { display: flex; gap: 0.5rem; }
.info-label { color: #888; min-width: 120px; }
.info-value { color: #222; font-weight: 500; }
[data-testid="stDownloadButton"] button {
    background: #1f5c2e !important; color: white !important;
    border-radius: 8px !important; font-weight: 600 !important;
    padding: 0.6rem 1.5rem !important; border: none !important;
}
[data-testid="stDownloadButton"] button:hover { background: #174d25 !important; }
.stSpinner > div { border-top-color: #1f5c2e !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
SHIPPING_PER_PLANT = 40
GREEN  = "1F5C2E"
LGREEN = "E8F5E9"
WHITE  = "FFFFFF"
GREY   = "F5F5F5"

IMG_ROW_HEIGHT = 175   # row height in points for product rows
IMG_COL_WIDTH  = 30.5  # column width in chars for image column

HEADER_FIELDS = [
    ("Order ID",          "Order ID"),
    ("Total Order Value", "Total Order Value"),
    ("Shipping cost",     "Shipping cost"),
    ("Total Tax Amount",  "Total Tax Amount"),
    ("Customer Name",     "Customer Name"),
    ("Customer Phone",    "Customer Phone"),
    ("City",              "City"),
    ("State",             "State"),
    ("Country",           "Country"),
    ("Pincode",           "Pincode"),
    ("Complete address",  "Complete address"),
]

# ── Parser ────────────────────────────────────────────────────────────────────

def parse_xlsx(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    kv = {}
    product_header_row = None
    col_idx = {}

    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]

        # Detect the product table header row
        if "Product Name" in cells and ("Product SKU" in cells or "Quantity" in cells):
            product_header_row = i
            # Record actual column index for each field we need
            for field in ("Product Name", "Product SKU", "Product Price",
                          "Discounted Price", "Quantity"):
                if field in cells:
                    col_idx[field] = cells.index(field)
            break

        # Accumulate key-value header info
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 2:
            kv[str(non_empty[0]).strip()] = non_empty[1]
        elif len(non_empty) == 1:
            kv[str(non_empty[0]).strip()] = None

    if product_header_row is None:
        raise ValueError("Could not find product table (looking for 'Product Name' header row).")

    # Resolve column indices with fallbacks
    name_col  = col_idx.get("Product Name")
    sku_col   = col_idx.get("Product SKU")
    price_col = col_idx.get("Product Price")
    qty_col   = col_idx.get("Quantity")

    if name_col is None:
        raise ValueError("Could not find 'Product Name' column in product table.")

    # Extract embedded images keyed by row index (0-based)
    img_by_row = {}
    for img in ws._images:
        try:
            row_idx = img.anchor._from.row
            img.ref.seek(0)
            data = img.ref.read()
            img_by_row[row_idx] = (data, img.format or "jpeg")
        except Exception:
            pass

    products = []
    for abs_row, row in enumerate(rows[product_header_row + 1:], product_header_row + 1):
        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        name  = cell(name_col)
        sku   = cell(sku_col)
        price = cell(price_col)
        qty   = cell(qty_col)

        # Stop at first completely empty row
        if name is None and sku is None:
            break

        name  = str(name).strip() if name else ""
        sku   = str(sku).strip()  if sku  else ""

        try:
            price = float(str(price).replace(",", "").strip()) if price not in (None, "", "None") else 0
        except (ValueError, TypeError):
            price = 0

        try:
            qty = int(float(str(qty).strip())) if qty not in (None, "", "None") else 1
        except (ValueError, TypeError):
            qty = 1

        img_data = img_by_row.get(abs_row)
        products.append({"name": name, "sku": sku, "price": price, "qty": qty, "img": img_data})

    return {"kv": kv, "products": products}


def parse_pdf(file_bytes: bytes) -> dict:
    import pdfplumber

    all_lines = []
    product_images = []  # ordered list of (bytes, "jpeg") matching product order

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            all_lines.extend((page.extract_text() or "").splitlines())

            # Product images sit in the left column (x0 ≈ 53); logo has x0 ≈ 33 — skip it
            page_imgs = sorted(
                [img for img in page.images if img["x0"] > 50],
                key=lambda x: x["top"],
            )
            for img in page_imgs:
                try:
                    product_images.append((img["stream"].get_data(), "jpeg"))
                except Exception:
                    product_images.append(None)

    full_text = "\n".join(all_lines)
    kv = {}
    m = re.search(r'Order ID[:\s]+(\S+)', full_text)
    if m:
        kv["Order ID"] = m.group(1)
    m = re.search(r'₹\s*([\d,]+)\s*\n', full_text)
    if m:
        kv["Total Order Value"] = m.group(1).replace(",", "")

    # Extract customer details if a CUSTOMER DETAILS section exists
    cust_idx = next((i for i, l in enumerate(all_lines) if "CUSTOMER DETAILS" in l), None)
    prod_hdr_idx = next((i for i, l in enumerate(all_lines) if "No. Product Item" in l), None)
    if cust_idx is not None and prod_hdr_idx is not None:
        section = all_lines[cust_idx + 1 : prod_hdr_idx]
        # Customer name: first non-empty line that isn't pricing noise
        noise_re = re.compile(r'(TOTAL|Estimate|₹|products|\d{4}$|^India$)', re.I)
        phone_re = re.compile(r'^\+\d')
        name_candidates = [l.strip() for l in section
                           if l.strip() and not phone_re.match(l.strip()) and not noise_re.search(l.strip())]
        kv["Customer Name"] = name_candidates[0] if name_candidates else ""

        # Phone
        phone = next((l.strip() for l in section if phone_re.match(l.strip())), "")
        kv["Customer Phone"] = phone

        # City / State / Pincode — look for "City, State - 6digits"
        city_state_re = re.compile(r'^(.+?),\s*(.+?)\s*-\s*(\d{6})')
        for l in section:
            cm = city_state_re.match(l.strip())
            if cm:
                kv["City"]    = cm.group(1).strip()
                kv["State"]   = cm.group(2).strip()
                kv["Pincode"] = cm.group(3).strip()
                break

        # Complete address: join cleaned address lines (skip name, phone, noise-only)
        addr_lines = []
        skip_name = kv.get("Customer Name", "")
        for l in section:
            l = l.strip()
            if not l or l == skip_name or phone_re.match(l) or l in ("India", "Estimate"):
                continue
            if re.match(r'^Order ID', l, re.I):
                continue
            l = re.sub(r'\s*TOTAL\s*$', '', l)
            l = re.sub(r'\s*₹\s*[\d,]+', '', l)
            l = re.sub(r'\d+\s*products', '', l).strip().rstrip(',')
            if l:
                addr_lines.append(l)
        kv["Complete address"] = ", ".join(addr_lines)

    product_re = re.compile(r'^(\d+)\s+(.+?)\s+(\d+)\s+₹\s*([\d,]+)\s+₹\s*([\d,]+)\s*$')
    sku_re     = re.compile(r'^SKU\s*:\s*(\S+)')

    products = []
    for i, line in enumerate(all_lines):
        pm = product_re.match(line.strip())
        if pm:
            _, name, qty, price, _ = pm.groups()
            sku = ""
            if i + 1 < len(all_lines):
                sm = sku_re.match(all_lines[i + 1].strip())
                if sm:
                    sku = sm.group(1)
            idx = len(products)
            products.append({
                "name":  name.strip(),
                "sku":   sku,
                "price": float(price.replace(",", "")),
                "qty":   int(qty),
                "img":   product_images[idx] if idx < len(product_images) else None,
            })

    if not products:
        raise ValueError("No products found in PDF. Is this a Quicksell estimate PDF?")

    return {"kv": kv, "products": products}


def img_to_b64(data: bytes, fmt: str) -> str:
    mime = "image/jpeg" if fmt.lower() in ("jpg", "jpeg") else f"image/{fmt.lower()}"
    return f"data:{mime};base64,{base64.b64encode(data).decode()}"


# ── Output xlsx builder ───────────────────────────────────────────────────────

def build_xlsx(data: dict, shipping: int) -> bytes:
    kv       = data["kv"]
    products = data["products"]

    thin = Side(style="thin", color="1F5C2E")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fl(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Sheet"

    # Col widths: Sr, Image, Name, SKU, Price, Qty, Amount
    col_widths = [6, IMG_COL_WIDTH, 42, 18, 13, 10, 13]
    for col, w in zip(range(1, 8), col_widths):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1
    for label, key in HEADER_FIELDS:
        val = kv.get(key, "")
        c = ws.cell(row, 2, label)
        c.font = Font(name="Arial", bold=True, size=10)
        ws.cell(row, 3, val).font = Font(name="Arial", size=10)
        row += 1
    row += 1  # blank spacer

    # Table header
    headers = ["Sr No", "Product Image", "Product Name", "Product SKU",
               "Product Price", "Quantity", "Amount"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill      = fl(GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
    ws.row_dimensions[row].height = 22
    row += 1

    product_start = row

    for i, p in enumerate(products, 1):
        row_fill = fl(GREY) if i % 2 == 0 else fl(WHITE)
        ws.row_dimensions[row].height = IMG_ROW_HEIGHT

        vals   = [i, "", p["name"], p["sku"], p["price"], p["qty"], f"=E{row}*F{row}"]
        aligns = ["center", "center", "left", "left", "center", "center", "center"]
        for col, (val, ha) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row, col, val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=(col == 3))
            c.fill      = row_fill
            c.border    = bdr

        # Embed image: move and size with cells, inset by ~8px on each side
        if p["img"]:
            try:
                img_data, fmt = p["img"]
                pil = PILImage.open(io.BytesIO(img_data))
                pil = pil.convert("RGB")
                pil.thumbnail((200, 200), PILImage.LANCZOS)

                buf = io.BytesIO()
                pil.save(buf, format="JPEG", quality=95)
                buf.seek(0)

                xl_img = XLImage(buf)

                # TwoCellAnchor: image moves and resizes with cell, 8px inset
                EMU_PER_PX = 9525
                pad = 8 * EMU_PER_PX
                r = row - 1  # 0-based row
                c = 1        # column B (0-based)
                _from = AnchorMarker(col=c,     colOff=pad,  row=r,     rowOff=pad)
                _to   = AnchorMarker(col=c + 1, colOff=-pad, row=r + 1, rowOff=-pad)
                anchor = TwoCellAnchor(editAs="twoCell")
                anchor._from = _from
                anchor.to    = _to
                xl_img.anchor = anchor
                ws.add_image(xl_img)
            except Exception:
                pass

        row += 1

    product_end = row - 1

    # Totals rows
    def total_row(label, qty_val, amt_val, bold=False, bg=None):
        nonlocal row
        for col in range(1, 8):
            c = ws.cell(row, col)
            c.border = bdr
            c.font   = Font(name="Arial", bold=bold, size=10,
                            color=WHITE if bg == GREEN else "000000")
            if bg:
                c.fill = fl(bg)
        ws.cell(row, 3, label).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 3).font = Font(name="Arial", bold=bold, size=10,
                                    color=WHITE if bg == GREEN else "000000")
        if qty_val is not None:
            ws.cell(row, 6, qty_val).alignment = Alignment(horizontal="center", vertical="center")
        if amt_val is not None:
            ws.cell(row, 7, amt_val).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

    subtotal_row = row
    total_row("Subtotal",
              f"=SUM(F{product_start}:F{product_end})",
              f"=SUM(G{product_start}:G{product_end})",
              bold=True, bg=LGREEN)
    total_row(f"Shipping & Packing (₹{SHIPPING_PER_PLANT} × plants)", None, shipping,
              bold=True)
    total_row("Final Amount", None, f"=G{row-2}+G{row-1}", bold=True, bg=GREEN)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ── Auth ─────────────────────────────────────────────────────────────────────

def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <div style="max-width:400px; margin: 8rem auto; text-align:center;">
      <div style="font-size:3rem; margin-bottom:1rem;">🌿</div>
      <h2 style="font-family:'DM Serif Display',serif; color:#1f5c2e; margin-bottom:0.25rem;">
        Order Sheet Generator
      </h2>
      <p style="color:#888; margin-bottom:2rem; font-size:0.95rem;">
        Chhajed Garden · Sanjay Nursery
      </p>
    </div>
    """, unsafe_allow_html=True)

    col = st.columns([1, 2, 1])[1]
    with col:
        pwd = st.text_input("Password", type="password", placeholder="Enter password…", label_visibility="collapsed")
        if pwd:
            if pwd == st.secrets.get("APP_PASSWORD", ""):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False

if not check_password():
    st.stop()

# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
  <div style="font-size:2.8rem">🌿</div>
  <div>
    <h1>Order Sheet Generator</h1>
    <p>Chhajed Garden · Sanjay Nursery — Upload a Quicksell export to generate a clean order sheet</p>
  </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Upload Quicksell order export (.xlsx or .pdf)",
    type=["xlsx", "pdf"],
    label_visibility="collapsed",
    help="Upload a Quicksell .xlsx export or .pdf estimate"
)

if uploaded:
    with st.spinner("Reading order..."):
        try:
            file_bytes = uploaded.read()
            if uploaded.name.lower().endswith(".pdf"):
                data = parse_pdf(file_bytes)
            else:
                data = parse_xlsx(file_bytes)
        except Exception as e:
            st.error(f"Could not read file: {e}")
            st.stop()

    kv       = data["kv"]
    products = data["products"]
    total_qty = sum(p["qty"] for p in products)
    shipping  = total_qty * SHIPPING_PER_PLANT
    subtotal  = sum(p["price"] * p["qty"] for p in products)
    final     = subtotal + shipping

    # Order info card
    st.markdown(f"""
    <div class="order-card">
      <h3>📦 Order {kv.get('Order ID','—')}</h3>
      <div class="info-grid">
        <div class="info-row"><span class="info-label">Customer</span><span class="info-value">{kv.get('Customer Name','—')}</span></div>
        <div class="info-row"><span class="info-label">Phone</span><span class="info-value">{kv.get('Customer Phone','—')}</span></div>
        <div class="info-row"><span class="info-label">City</span><span class="info-value">{kv.get('City','—')}, {kv.get('State','—')}</span></div>
        <div class="info-row"><span class="info-label">Pincode</span><span class="info-value">{kv.get('Pincode','—')}</span></div>
        <div class="info-row" style="grid-column:1/-1"><span class="info-label">Address</span><span class="info-value">{kv.get('Complete address','—')}</span></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Product table
    rows_html = ""
    for i, p in enumerate(products, 1):
        if p["img"]:
            img_b64 = img_to_b64(p["img"][0], p["img"][1])
            img_tag = f'<img src="{img_b64}" alt="{p["name"]}">'
        else:
            img_tag = '<div style="width:64px;height:64px;background:#f0f0f0;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:#bbb">🌱</div>'

        amount = p["price"] * p["qty"]
        rows_html += f"""
        <tr>
          <td>{i}</td>
          <td class="img-cell">{img_tag}</td>
          <td>{p['name']}</td>
          <td class="sku">{p['sku']}</td>
          <td>₹{p['price']:,.0f}</td>
          <td>{p['qty']}</td>
          <td class="amt">₹{amount:,.0f}</td>
        </tr>"""

    st.markdown(f"""
    <table class="product-table">
      <thead>
        <tr>
          <th>#</th><th>Image</th><th>Product Name</th>
          <th>SKU</th><th>Price</th><th>Qty</th><th>Amount</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)

    # Totals
    st.markdown(f"""
    <div class="totals-block">
      <div class="totals-row">
        <span class="totals-label">Subtotal ({len(products)} items)</span>
        <span class="totals-value">₹{subtotal:,.0f}</span>
      </div>
      <div class="totals-row">
        <span class="totals-label">Shipping & Packing (₹{SHIPPING_PER_PLANT} × {total_qty} plants)</span>
        <span class="totals-value">₹{shipping:,.0f}</span>
      </div>
      <div class="totals-row final">
        <span class="totals-label">Final Amount</span>
        <span class="totals-value">₹{final:,.0f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    with st.spinner("Building Excel file..."):
        xlsx_bytes = build_xlsx(data, shipping)

    customer_slug = re.sub(r'[^\w]', '_', str(kv.get('Customer Name', 'order')).lower())
    st.download_button(
        label="⬇️  Download Order Sheet (.xlsx)",
        data=xlsx_bytes,
        file_name=f"{customer_slug}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.markdown("""
    <div style="text-align:center; padding: 3rem 1rem; color: #999;">
      <div style="font-size:3rem; margin-bottom:1rem">📂</div>
      <p style="font-size:1rem;">Upload a Quicksell .xlsx export above to get started</p>
    </div>
    """, unsafe_allow_html=True)
